from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Form, BackgroundTasks
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from docling.document_converter import DocumentConverter
from docling.chunking import HybridChunker
from docling_core.transforms.chunker.tokenizer.huggingface import HuggingFaceTokenizer
from transformers import AutoTokenizer
import tempfile
import os
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Literal, Optional, List, Dict
import logging
import mimetypes
import uuid
from datetime import datetime
import asyncio

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 작업 상태 저장소 (실제 프로덕션에서는 Redis나 DB 사용 권장)
job_store: Dict[str, dict] = {}

app = FastAPI(
    title="Docling Document Converter",
    description="Docling HybridChunker와 Contextualize를 활용한 다양한 파일 형식을 Markdown으로 변환하는 API",
    version="1.0.0"
)

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Docling 컨버터 및 HybridChunker 초기화
doc_converter = DocumentConverter()

# HuggingFace 토크나이저 설정 (한글 최적화)
EMBED_MODEL_ID = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"  # 한글 처리 3배 효율적!
MAX_TOKENS = 512  # 청크 크기 제한

tokenizer = HuggingFaceTokenizer(
    tokenizer=AutoTokenizer.from_pretrained(EMBED_MODEL_ID),
    max_tokens=MAX_TOKENS,
)

# HybridChunker 초기화
chunker = HybridChunker(
    tokenizer=tokenizer,
    merge_peers=True,  # 인접한 청크 병합 옵션
)

# 지원되는 파일 확장자 (Docling 지원 형식)
SUPPORTED_EXTENSIONS = {
    '.pdf': 'PDF 문서',
    '.docx': 'Word 문서',
    '.xlsx': 'Excel 스프레드시트',
    '.pptx': 'PowerPoint 프레젠테이션',
    '.html': 'HTML 파일',
    '.htm': 'HTML 파일',
    '.md': 'Markdown 파일',
    '.txt': '텍스트 파일',
    '.json': 'JSON 파일',
    '.xml': 'XML 파일',
    '.csv': 'CSV 파일',
    '.jpg': 'JPEG 이미지',
    '.jpeg': 'JPEG 이미지',
    '.png': 'PNG 이미지',
    '.gif': 'GIF 이미지',
    '.bmp': 'BMP 이미지',
}

@app.get("/")
async def root():
    """API 상태 및 지원 파일 형식 확인"""
    return {
        "message": "Docling Document Converter API",
        "status": "running",
        "supported_formats": SUPPORTED_EXTENSIONS,
        "chunking": {
            "engine": "HybridChunker",
            "tokenizer": EMBED_MODEL_ID,
            "max_tokens": MAX_TOKENS,
            "contextualize": True
        },
        "endpoints": {
            "convert": "/convert - 파일을 Markdown으로 변환",
            "convert-chunked": "/convert-chunked - 파일을 청크별로 변환 (동기)",
            "convert-chunked-async": "/convert-chunked-async - 파일을 청크별로 변환 (비동기, 대용량 파일용)",
            "convert-multiple": "/convert-multiple - 여러 파일을 일괄 변환",
            "job-status": "/job/{job_id} - 비동기 작업 상태 조회",
            "jobs": "/jobs - 모든 작업 목록 조회",
            "delete-job": "DELETE /job/{job_id} - 작업 삭제 및 정리",
            "health": "/health - 서버 상태 확인"
        },
        "async_workflow": {
            "description": "대용량 파일 처리 시 타임아웃 방지를 위한 비동기 워크플로우",
            "steps": [
                "1. POST /convert-chunked-async - 파일 업로드 및 job_id 수신",
                "2. GET /job/{job_id} - 작업 상태 polling (status: queued/processing/completed/failed)",
                "3. status가 'completed'면 result에서 청크 데이터 획득",
                "4. DELETE /job/{job_id} - 작업 정리 (선택사항)"
            ]
        }
    }

@app.get("/health")
async def health_check():
    """서버 상태 확인"""
    return {
        "status": "healthy", 
        "library": "Docling with HybridChunker", 
        "supported_files": len(SUPPORTED_EXTENSIONS),
        "tokenizer": EMBED_MODEL_ID,
        "max_tokens": MAX_TOKENS,
        "active_jobs": len(job_store)
    }

@app.post("/convert")
async def convert_file(
    file: UploadFile = File(..., description="변환할 파일"),
    output_filename: Optional[str] = Query(None, description="출력 파일명 (확장자 제외)"),
    include_metadata: bool = Query(False, description="파일 메타데이터 포함"),
    use_chunking: bool = Query(False, description="HybridChunker를 사용한 청킹 적용"),
    contextualize: bool = Query(True, description="청크 컨텍스트 강화 적용")
):
    """
    단일 파일을 Markdown으로 변환합니다.
    
    HybridChunker를 사용하여 문서를 의미있는 청크로 분할하고,
    contextualize 기능으로 각 청크의 컨텍스트를 강화할 수 있습니다.
    
    지원 형식:
    - 문서: PDF, Word (.docx), PowerPoint (.pptx), Excel (.xlsx)
    - 웹: HTML, Markdown, CSV, JSON, XML
    - 이미지: JPG, PNG, GIF, BMP
    - 텍스트: TXT
    """
    
    # 파일 확장자 검증
    file_extension = Path(file.filename).suffix.lower()
    if file_extension not in SUPPORTED_EXTENSIONS:
        supported_list = ', '.join(SUPPORTED_EXTENSIONS.keys())
        raise HTTPException(
            status_code=400, 
            detail=f"지원되지 않는 파일 형식입니다. 지원 형식: {supported_list}"
        )
    
    # 임시 디렉토리에서 처리
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            # 업로드된 파일 저장
            input_file_path = os.path.join(temp_dir, file.filename)
            with open(input_file_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)
            
            logger.info(f"파일 저장 완료: {input_file_path} ({len(content)} bytes)")
            
            # Docling으로 변환
            logger.info("Docling 문서 변환 시작")
            result = doc_converter.convert(input_file_path)
            doc = result.document
            logger.info("Docling 문서 변환 완료")
            
            # 청킹 적용 여부에 따른 처리
            if use_chunking:
                logger.info("HybridChunker로 청킹 시작")
                chunks = list(chunker.chunk(dl_doc=doc))
                logger.info(f"청킹 완료: {len(chunks)}개 청크 생성")
                
                # 각 청크를 처리하여 Markdown 생성
                markdown_parts = []
                for i, chunk in enumerate(chunks):
                    if contextualize:
                        # contextualize로 컨텍스트 강화된 텍스트 생성
                        enriched_text = chunker.contextualize(chunk=chunk)
                        markdown_parts.append(f"## 청크 {i+1}\n\n{enriched_text}\n")
                    else:
                        # 기본 청크 텍스트 사용
                        markdown_parts.append(f"## 청크 {i+1}\n\n{chunk.text}\n")
                
                final_content = "\n".join(markdown_parts)
            else:
                # 기본 Markdown 변환 (청킹 없음)
                final_content = doc.export_to_markdown()
            
            # 출력 파일명 설정
            if not output_filename:
                output_filename = Path(file.filename).stem
            
            # Markdown 파일로 응답
            return Response(
                content=final_content,
                media_type="text/markdown",
                headers={
                    "Content-Disposition": f"attachment; filename={output_filename}.md"
                }
            )
            
        except Exception as e:
            logger.error(f"변환 중 오류 발생: {str(e)}")
            raise HTTPException(status_code=500, detail=f"파일 변환 중 오류가 발생했습니다: {str(e)}")

@app.post("/convert-chunked")
async def convert_file_chunked(
    file: UploadFile = File(..., description="변환할 파일"),
    include_metadata: bool = Query(False, description="청크 메타데이터 포함"),
    contextualize: bool = Query(True, description="청크 컨텍스트 강화 적용"),
    max_tokens: Optional[int] = Query(None, description="최대 토큰 수 (기본값: 512)")
):
    """
    파일을 HybridChunker로 청킹하여 JSON 형태로 응답합니다.
    각 청크의 상세 정보와 컨텍스트 강화된 텍스트를 제공합니다.
    """
    
    file_extension = Path(file.filename).suffix.lower()
    if file_extension not in SUPPORTED_EXTENSIONS:
        supported_list = ', '.join(SUPPORTED_EXTENSIONS.keys())
        raise HTTPException(
            status_code=400, 
            detail=f"지원되지 않는 파일 형식입니다. 지원 형식: {supported_list}"
        )
    
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            # 파일 저장
            input_file_path = os.path.join(temp_dir, file.filename)
            with open(input_file_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)
            
            # 사용자 정의 토큰 수가 있다면 새로운 chunker 생성
            if max_tokens and max_tokens != MAX_TOKENS:
                custom_tokenizer = HuggingFaceTokenizer(
                    tokenizer=AutoTokenizer.from_pretrained(EMBED_MODEL_ID),
                    max_tokens=max_tokens,
                )
                custom_chunker = HybridChunker(
                    tokenizer=custom_tokenizer,
                    merge_peers=True,
                )
                active_chunker = custom_chunker
            else:
                active_chunker = chunker
            
            # 문서 변환
            result = doc_converter.convert(input_file_path)
            doc = result.document
            
            # Excel 파일인 경우 시트 정보 직접 추출
            excel_sheets = []
            if file_extension == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(input_file_path, read_only=True)
                    excel_sheets = wb.sheetnames
                    wb.close()
                    logger.info(f"Excel 시트 감지: {excel_sheets}")
                except Exception as e:
                    logger.warning(f"Excel 시트 이름 추출 실패: {e}")
            
            # HybridChunker로 청킹
            chunks = list(active_chunker.chunk(dl_doc=doc))
            
            # 청크 정보 수집
            chunk_data = []
            for i, chunk in enumerate(chunks):
                chunk_info = {
                    "chunk_id": i + 1,
                    "text": chunk.text,
                    "text_length": len(chunk.text),
                    "token_count": tokenizer.count_tokens(chunk.text) if hasattr(tokenizer, 'count_tokens') else None
                }
                
                # 컨텍스트 강화 적용
                if contextualize:
                    enriched_text = active_chunker.contextualize(chunk=chunk)
                    chunk_info["contextualized_text"] = enriched_text
                    chunk_info["contextualized_length"] = len(enriched_text)
                
                # 페이지 정보 및 확장된 메타데이터 추가 (항상 활성화)
                # 기본 메타데이터
                if hasattr(chunk, 'meta'):
                    chunk_info["metadata"] = chunk.meta.export_json_dict() if hasattr(chunk.meta, 'export_json_dict') else str(chunk.meta)
                
                # 디버그: 청크 구조 확인
                chunk_info["debug_chunk_attrs"] = [attr for attr in dir(chunk) if not attr.startswith('_')]
                if hasattr(chunk, 'meta'):
                    chunk_info["debug_meta_attrs"] = [attr for attr in dir(chunk.meta) if not attr.startswith('_')]
                
                # 페이지 정보 추출 - 파일 타입별 처리
                page_numbers = set()
                bbox_data = []
                sheet_names = set()  # Excel 시트 이름
                
                # PDF 등 페이지 기반 문서
                if hasattr(chunk, 'meta') and hasattr(chunk.meta, 'doc_items'):
                    for doc_item in chunk.meta.doc_items:
                        if hasattr(doc_item, 'prov') and isinstance(doc_item.prov, list):
                            for prov in doc_item.prov:
                                # page_no가 있으면 사용 (PDF)
                                if hasattr(prov, 'page_no') and prov.page_no > 0:
                                    page_numbers.add(prov.page_no)
                                    
                                    bbox_info = {
                                        "page": prov.page_no,
                                        "label": getattr(doc_item, 'label', 'unknown')
                                    }
                                    
                                    # bbox 정보 추가
                                    if hasattr(prov, 'bbox'):
                                        bbox_info["bbox"] = {
                                            "l": getattr(prov.bbox, 'l', 0),
                                            "t": getattr(prov.bbox, 't', 0), 
                                            "r": getattr(prov.bbox, 'r', 0),
                                            "b": getattr(prov.bbox, 'b', 0),
                                            "coord_origin": getattr(prov.bbox, 'coord_origin', 'BOTTOMLEFT')
                                        }
                                    
                                    # charspan 정보 추가
                                    if hasattr(prov, 'charspan'):
                                        bbox_info["charspan"] = prov.charspan
                                    
                                    bbox_data.append(bbox_info)
                                
                                # Excel 시트 이름 추출 시도
                                if hasattr(prov, 'sheet_name'):
                                    sheet_names.add(prov.sheet_name)
                                elif hasattr(prov, 'page_name'):
                                    sheet_names.add(prov.page_name)
                
                # 페이지 정보 설정
                # 파일 타입별로 다르게 처리
                if file_extension == '.xlsx':
                    # Excel: page_no는 시트 순서, 별도로 sheet 필드 사용
                    chunk_info["page_info"] = []
                    chunk_info["pages"] = []  # Excel은 페이지 없음
                    
                    # 시트 정보 추가
                    if sheet_names:
                        chunk_info["sheet_names"] = sorted(list(sheet_names))
                        chunk_info["sheet"] = list(sheet_names)[0] if len(sheet_names) == 1 else sorted(list(sheet_names))
                    elif excel_sheets and page_numbers:
                        # page_no로 시트 매칭 (page_no 1 = 첫번째 시트)
                        matched_sheets = []
                        for page_no in sorted(page_numbers):
                            if page_no <= len(excel_sheets):
                                matched_sheets.append(excel_sheets[page_no - 1])
                        if matched_sheets:
                            chunk_info["sheet_names"] = matched_sheets
                            chunk_info["sheet"] = matched_sheets[0] if len(matched_sheets) == 1 else matched_sheets
                            chunk_info["sheet_index"] = sorted(list(page_numbers))  # 시트 순서
                    
                elif page_numbers:
                    # PDF 등: 페이지 정보
                    chunk_info["page_info"] = sorted(list(page_numbers))
                    chunk_info["pages"] = sorted(list(page_numbers))
                else:
                    # DOCX 등: 섹션 정보
                    chunk_info["page_info"] = []
                    chunk_info["pages"] = []
                    chunk_info["section_index"] = i + 1  # 청크 순서
                    
                chunk_info["bbox_info"] = bbox_data
                
                chunk_data.append(chunk_info)
            
            # JSON 응답 준비
            response_data = {
                "success": True,
                "filename": file.filename,
                "file_type": SUPPORTED_EXTENSIONS.get(file_extension, "알 수 없음"),
                "total_chunks": len(chunks),
                "chunking_config": {
                    "tokenizer": EMBED_MODEL_ID,
                    "max_tokens": max_tokens or MAX_TOKENS,
                    "merge_peers": True,
                    "contextualize": contextualize
                },
                "chunks": chunk_data
            }
            
            if include_metadata:
                # 문서 전체 페이지 정보 추출
                total_pages = 0
                if hasattr(doc, 'pages') and doc.pages:
                    total_pages = len(doc.pages)
                elif hasattr(doc, 'main_text') and hasattr(doc.main_text, 'page'):
                    # 문서의 모든 아이템에서 최대 페이지 번호 찾기
                    max_page = 0
                    for item in doc.main_text:
                        if hasattr(item, 'prov') and hasattr(item.prov, 'page'):
                            max_page = max(max_page, item.prov.page)
                    total_pages = max_page
                
                response_data["file_metadata"] = {
                    "original_file_size": len(content),
                    "mime_type": mimetypes.guess_type(file.filename)[0],
                    "extension": file_extension,
                    "total_pages": total_pages if total_pages > 0 else None
                }
            
            return response_data
            
        except Exception as e:
            logger.error(f"변환 중 오류 발생: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "filename": file.filename
            }

# ==================== 비동기 작업 엔드포인트 (n8n 타임아웃 방지) ====================

def process_chunking_job(
    job_id: str,
    file_path: str,
    filename: str,
    file_extension: str,
    include_metadata: bool,
    contextualize: bool,
    max_tokens: Optional[int]
):
    """백그라운드에서 청킹 작업 처리 (동기 함수로 실행됨)"""
    logger.info(f"🚀 [Job {job_id}] 백그라운드 작업 시작됨! 파일: {filename}")
    try:
        job_store[job_id]["status"] = "processing"
        job_store[job_id]["progress"] = 10
        logger.info(f"[Job {job_id}] 청킹 작업 시작: {filename}")
        
        # 문서 변환
        job_store[job_id]["progress"] = 30
        job_store[job_id]["message"] = "문서 변환 중..."
        result = doc_converter.convert(file_path)
        doc = result.document
        logger.info(f"[Job {job_id}] 문서 변환 완료")
        
        # 사용자 정의 토큰 수가 있다면 새로운 chunker 생성
        job_store[job_id]["progress"] = 40
        if max_tokens and max_tokens != MAX_TOKENS:
            custom_tokenizer = HuggingFaceTokenizer(
                tokenizer=AutoTokenizer.from_pretrained(EMBED_MODEL_ID),
                max_tokens=max_tokens,
            )
            custom_chunker = HybridChunker(
                tokenizer=custom_tokenizer,
                merge_peers=True,
            )
            active_chunker = custom_chunker
        else:
            active_chunker = chunker
        
        # Excel 파일인 경우 시트 정보 직접 추출
        excel_sheets = []
        if file_extension == '.xlsx':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                excel_sheets = wb.sheetnames
                wb.close()
                logger.info(f"[Job {job_id}] Excel 시트 감지: {excel_sheets}")
            except Exception as e:
                logger.warning(f"[Job {job_id}] Excel 시트 이름 추출 실패: {e}")
        
        # HybridChunker로 청킹
        job_store[job_id]["progress"] = 60
        job_store[job_id]["message"] = "청킹 중..."
        chunks = list(active_chunker.chunk(dl_doc=doc))
        logger.info(f"[Job {job_id}] 청킹 완료: {len(chunks)}개 청크 생성")
        
        # 청크 정보 수집
        job_store[job_id]["progress"] = 80
        job_store[job_id]["message"] = "청크 메타데이터 처리 중..."
        chunk_data = []
        for i, chunk in enumerate(chunks):
            chunk_info = {
                "chunk_id": i + 1,
                "text": chunk.text,
                "text_length": len(chunk.text),
                "token_count": tokenizer.count_tokens(chunk.text) if hasattr(tokenizer, 'count_tokens') else None
            }
            
            # 컨텍스트 강화 적용
            if contextualize:
                enriched_text = active_chunker.contextualize(chunk=chunk)
                chunk_info["contextualized_text"] = enriched_text
                chunk_info["contextualized_length"] = len(enriched_text)
            
            # 메타데이터 처리 (완전한 버전 - 페이지 및 시트 정보 포함)
            if hasattr(chunk, 'meta'):
                chunk_info["metadata"] = chunk.meta.export_json_dict() if hasattr(chunk.meta, 'export_json_dict') else str(chunk.meta)
            
            # 페이지 및 시트 정보 추출
            page_numbers = set()
            bbox_data = []
            sheet_names = set()
            
            if hasattr(chunk, 'meta') and hasattr(chunk.meta, 'doc_items'):
                for doc_item in chunk.meta.doc_items:
                    if hasattr(doc_item, 'prov') and isinstance(doc_item.prov, list):
                        for prov in doc_item.prov:
                            # PDF 페이지 정보
                            if hasattr(prov, 'page_no') and prov.page_no > 0:
                                page_numbers.add(prov.page_no)
                                
                                bbox_info = {
                                    "page": prov.page_no,
                                    "label": getattr(doc_item, 'label', 'unknown')
                                }
                                
                                # bbox 정보 추가
                                if hasattr(prov, 'bbox'):
                                    bbox_info["bbox"] = {
                                        "l": getattr(prov.bbox, 'l', 0),
                                        "t": getattr(prov.bbox, 't', 0),
                                        "r": getattr(prov.bbox, 'r', 0),
                                        "b": getattr(prov.bbox, 'b', 0),
                                        "coord_origin": getattr(prov.bbox, 'coord_origin', 'BOTTOMLEFT')
                                    }
                                
                                # charspan 정보 추가
                                if hasattr(prov, 'charspan'):
                                    bbox_info["charspan"] = prov.charspan
                                
                                bbox_data.append(bbox_info)
                            
                            # Excel 시트 정보
                            if hasattr(prov, 'sheet_name'):
                                sheet_names.add(prov.sheet_name)
                            elif hasattr(prov, 'page_name'):
                                sheet_names.add(prov.page_name)
            
            # 파일 타입별 정보 추가
            if file_extension == '.xlsx':
                # Excel: page_no는 시트 순서, 별도로 sheet 필드 사용
                chunk_info["page_info"] = []
                chunk_info["pages"] = []  # Excel은 페이지 없음
                
                # 시트 정보 추가
                if sheet_names:
                    chunk_info["sheet_names"] = sorted(list(sheet_names))
                    chunk_info["sheet"] = list(sheet_names)[0] if len(sheet_names) == 1 else sorted(list(sheet_names))
                elif excel_sheets and page_numbers:
                    # page_no로 시트 매칭 (page_no 1 = 첫번째 시트)
                    matched_sheets = []
                    for page_no in sorted(page_numbers):
                        if page_no <= len(excel_sheets):
                            matched_sheets.append(excel_sheets[page_no - 1])
                    if matched_sheets:
                        chunk_info["sheet_names"] = matched_sheets
                        chunk_info["sheet"] = matched_sheets[0] if len(matched_sheets) == 1 else matched_sheets
                        chunk_info["sheet_index"] = sorted(list(page_numbers))  # 시트 순서
                
            elif page_numbers:
                # PDF 등: 페이지 정보
                chunk_info["page_info"] = sorted(list(page_numbers))
                chunk_info["pages"] = sorted(list(page_numbers))
            else:
                # DOCX 등: 섹션 정보
                chunk_info["page_info"] = []
                chunk_info["pages"] = []
                chunk_info["section_index"] = i + 1  # 청크 순서
            
            # bbox 정보 추가
            chunk_info["bbox_info"] = bbox_data
            
            chunk_data.append(chunk_info)
        
        # Excel 시트 정보를 결과에 추가
        result_data = {
            "success": True,
            "filename": filename,
            "file_type": SUPPORTED_EXTENSIONS.get(file_extension, "알 수 없음"),
            "total_chunks": len(chunks),
            "chunking_config": {
                "tokenizer": EMBED_MODEL_ID,
                "max_tokens": max_tokens or MAX_TOKENS,
                "merge_peers": True,
                "contextualize": contextualize
            },
            "chunks": chunk_data
        }
        
        # Excel 파일인 경우 전체 시트 목록 추가
        if file_extension == '.xlsx' and excel_sheets:
            result_data["excel_sheets"] = excel_sheets
            result_data["total_sheets"] = len(excel_sheets)
        
        # 최종 결과 저장
        job_store[job_id]["progress"] = 100
        job_store[job_id]["status"] = "completed"
        job_store[job_id]["message"] = "처리 완료"
        job_store[job_id]["result"] = result_data
        job_store[job_id]["completed_at"] = datetime.now().isoformat()
        logger.info(f"✅ [Job {job_id}] 작업 완료!")
        
    except Exception as e:
        logger.error(f"❌ [Job {job_id}] 작업 실패: {str(e)}", exc_info=True)
        job_store[job_id]["status"] = "failed"
        job_store[job_id]["progress"] = 0
        job_store[job_id]["error"] = str(e)
        job_store[job_id]["failed_at"] = datetime.now().isoformat()
    
    finally:
        # 임시 파일 정리
        temp_dir = job_store[job_id].get("temp_dir")
        if temp_dir and os.path.exists(temp_dir):
            try:
                import shutil
                shutil.rmtree(temp_dir)
                logger.info(f"🗑️ [Job {job_id}] 임시 디렉토리 정리 완료: {temp_dir}")
            except Exception as cleanup_error:
                logger.warning(f"⚠️ [Job {job_id}] 임시 디렉토리 정리 실패: {cleanup_error}")

@app.post("/convert-chunked-async")
async def convert_file_chunked_async(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="변환할 파일"),
    include_metadata: bool = Query(False, description="청크 메타데이터 포함"),
    contextualize: bool = Query(True, description="청크 컨텍스트 강화 적용"),
    max_tokens: Optional[int] = Query(None, description="최대 토큰 수 (기본값: 512)")
):
    """
    비동기 청킹 작업을 시작하고 즉시 job_id를 반환합니다.
    대용량 파일 처리 시 타임아웃을 방지하기 위해 사용합니다.
    
    반환된 job_id로 /job/{job_id} 엔드포인트를 polling하여 상태를 확인할 수 있습니다.
    """
    logger.info(f"📥 [API 호출] POST /convert-chunked-async - 파일: {file.filename}")
    
    file_extension = Path(file.filename).suffix.lower()
    if file_extension not in SUPPORTED_EXTENSIONS:
        supported_list = ', '.join(SUPPORTED_EXTENSIONS.keys())
        logger.warning(f"❌ [API 호출] 지원되지 않는 파일 형식: {file_extension}")
        raise HTTPException(
            status_code=400, 
            detail=f"지원되지 않는 파일 형식입니다. 지원 형식: {supported_list}"
        )
    
    # Job ID 생성
    job_id = str(uuid.uuid4())
    logger.info(f"🆔 [API 호출] 생성된 Job ID: {job_id}")
    
    # 임시 파일로 저장
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Job 상태 초기화
        job_store[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "progress": 0,
            "message": "작업 대기 중...",
            "filename": file.filename,
            "file_size": len(content),
            "created_at": datetime.now().isoformat(),
            "temp_dir": temp_dir
        }
        
        # 백그라운드 작업 시작
        background_tasks.add_task(
            process_chunking_job,
            job_id=job_id,
            file_path=file_path,
            filename=file.filename,
            file_extension=file_extension,
            include_metadata=include_metadata,
            contextualize=contextualize,
            max_tokens=max_tokens
        )
        
        logger.info(f"[Job {job_id}] 비동기 작업 생성: {file.filename} ({len(content)} bytes)")
        
        return {
            "job_id": job_id,
            "status": "queued",
            "message": "작업이 시작되었습니다. /job/{job_id} 엔드포인트로 상태를 확인하세요.",
            "filename": file.filename,
            "file_size": len(content),
            "created_at": job_store[job_id]["created_at"]
        }
        
    except Exception as e:
        # 오류 발생 시 임시 디렉토리 정리
        if os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir)
        logger.error(f"작업 생성 실패: {str(e)}")
        raise HTTPException(status_code=500, detail=f"작업 생성 중 오류가 발생했습니다: {str(e)}")

@app.get("/job/{job_id}")
async def get_job_status(job_id: str):
    """
    작업 상태를 조회합니다.
    
    n8n에서 이 엔드포인트를 polling하여 작업 완료 여부를 확인할 수 있습니다.
    """
    logger.info(f"🔍 [API 호출] GET /job/{job_id} - 상태 조회 요청 받음")
    
    if job_id not in job_store:
        logger.warning(f"❌ [API 호출] GET /job/{job_id} - 작업 ID를 찾을 수 없음")
        raise HTTPException(status_code=404, detail=f"작업 ID '{job_id}'를 찾을 수 없습니다.")
    
    # 안전한 딕셔너리 읽기 (deep copy)
    import copy
    job = copy.deepcopy(job_store[job_id])
    
    logger.info(f"📊 [API 호출] GET /job/{job_id} - 현재 상태: {job['status']}, 진행률: {job.get('progress', 0)}%")
    
    response = {
        "job_id": job_id,
        "status": job["status"],
        "progress": job.get("progress", 0),
        "message": job.get("message", ""),
        "filename": job.get("filename"),
        "created_at": job.get("created_at")
    }
    
    if job["status"] == "completed":
        response["completed_at"] = job.get("completed_at")
        response["result"] = job.get("result")
        logger.info(f"✅ [API 호출] GET /job/{job_id} - 작업 완료 응답 반환")
    elif job["status"] == "failed":
        response["failed_at"] = job.get("failed_at")
        response["error"] = job.get("error")
        logger.error(f"❌ [API 호출] GET /job/{job_id} - 작업 실패 응답 반환: {job.get('error')}")
    else:
        logger.info(f"⏳ [API 호출] GET /job/{job_id} - 진행 중 응답 반환")
    
    return response

@app.delete("/job/{job_id}")
async def delete_job(job_id: str):
    """
    완료된 작업을 삭제하고 임시 파일을 정리합니다.
    """
    if job_id not in job_store:
        raise HTTPException(status_code=404, detail=f"작업 ID '{job_id}'를 찾을 수 없습니다.")
    
    job = job_store[job_id]
    
    # 임시 디렉토리 정리
    if "temp_dir" in job and os.path.exists(job["temp_dir"]):
        import shutil
        shutil.rmtree(job["temp_dir"])
        logger.info(f"[Job {job_id}] 임시 파일 정리 완료")
    
    # Job 삭제
    del job_store[job_id]
    logger.info(f"[Job {job_id}] 작업 삭제 완료")
    
    return {
        "message": f"작업 {job_id}가 삭제되었습니다.",
        "job_id": job_id
    }

@app.get("/jobs")
async def list_jobs():
    """
    모든 작업 목록을 조회합니다.
    """
    jobs = []
    for job_id, job in job_store.items():
        jobs.append({
            "job_id": job_id,
            "status": job["status"],
            "progress": job.get("progress", 0),
            "filename": job.get("filename"),
            "created_at": job.get("created_at")
        })
    
    return {
        "total_jobs": len(jobs),
        "jobs": jobs
    }

# ==================== 기존 엔드포인트 ====================

@app.post("/convert-multiple")
async def convert_multiple_files(
    files: List[UploadFile] = File(..., description="변환할 파일들"),
    output_format: Literal["zip", "json"] = Query("zip", description="출력 형식"),
    use_chunking: bool = Query(False, description="HybridChunker를 사용한 청킹 적용"),
    contextualize: bool = Query(True, description="청크 컨텍스트 강화 적용")
):
    """
    여러 파일을 일괄 변환합니다.
    
    - zip: 각 파일을 Markdown으로 변환하여 ZIP 파일로 반환
    - json: 모든 변환 결과를 JSON으로 반환
    """
    
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="한 번에 최대 10개 파일까지 처리 가능합니다.")
    
    results = []
    converted_files = {}
    
    with tempfile.TemporaryDirectory() as temp_dir:
        for file in files:
            try:
                file_extension = Path(file.filename).suffix.lower()
                
                # 지원되지 않는 파일은 스킵
                if file_extension not in SUPPORTED_EXTENSIONS:
                    results.append({
                        "filename": file.filename,
                        "success": False,
                        "error": "지원되지 않는 파일 형식"
                    })
                    continue
                
                # 파일 저장
                input_file_path = os.path.join(temp_dir, file.filename)
                with open(input_file_path, "wb") as buffer:
                    content = await file.read()
                    buffer.write(content)
                
                # Docling 변환
                result = doc_converter.convert(input_file_path)
                doc = result.document
                
                # 청킹 적용 여부에 따른 처리
                if use_chunking:
                    chunks = list(chunker.chunk(dl_doc=doc))
                    markdown_parts = []
                    for i, chunk in enumerate(chunks):
                        if contextualize:
                            enriched_text = chunker.contextualize(chunk=chunk)
                            markdown_parts.append(f"## 청크 {i+1}\n\n{enriched_text}\n")
                        else:
                            markdown_parts.append(f"## 청크 {i+1}\n\n{chunk.text}\n")
                    final_content = "\n".join(markdown_parts)
                else:
                    final_content = doc.export_to_markdown()
                
                # 결과 저장
                base_name = Path(file.filename).stem
                converted_files[f"{base_name}.md"] = final_content
                
                results.append({
                    "filename": file.filename,
                    "success": True,
                    "output_filename": f"{base_name}.md",
                    "content_length": len(final_content),
                    "chunks_count": len(list(chunker.chunk(dl_doc=doc))) if use_chunking else 1
                })
                
                logger.info(f"변환 완료: {file.filename}")
                
            except Exception as e:
                results.append({
                    "filename": file.filename,
                    "success": False,
                    "error": str(e)
                })
                logger.error(f"파일 변환 실패 ({file.filename}): {str(e)}")
    
    # JSON 형식으로 응답
    if output_format == "json":
        return {
            "total_files": len(files),
            "successful_conversions": len(converted_files),
            "chunking_applied": use_chunking,
            "contextualize_applied": contextualize,
            "results": results,
            "converted_content": {
                filename: content 
                for filename, content in converted_files.items()
            }
        }
    
    # ZIP 형식으로 응답
    if converted_files:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, content in converted_files.items():
                zip_file.writestr(filename, content)
        
        zip_buffer.seek(0)
        
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=docling_converted_files.zip"}
        )
    else:
        raise HTTPException(status_code=400, detail="변환 가능한 파일이 없습니다.")

@app.get("/supported-formats")
async def get_supported_formats():
    """지원되는 파일 형식 목록 반환"""
    return {
        "supported_extensions": SUPPORTED_EXTENSIONS,
        "total_formats": len(SUPPORTED_EXTENSIONS),
        "categories": {
            "documents": [".pdf", ".docx", ".pptx", ".xlsx"],
            "web": [".html", ".htm", ".md", ".csv", ".json", ".xml"],
            "images": [".jpg", ".jpeg", ".png", ".gif", ".bmp"],
            "text": [".txt"]
        },
        "chunking_info": {
            "engine": "HybridChunker",
            "tokenizer": EMBED_MODEL_ID,
            "default_max_tokens": MAX_TOKENS,
            "features": ["contextualize", "merge_peers", "hierarchical_chunking"]
        }
    }

@app.get("/chunking-info")
async def get_chunking_info():
    """HybridChunker 설정 정보 반환"""
    return {
        "chunker_type": "HybridChunker",
        "tokenizer": {
            "model": EMBED_MODEL_ID,
            "max_tokens": MAX_TOKENS,
            "type": "HuggingFaceTokenizer"
        },
        "features": {
            "contextualize": "청크의 컨텍스트를 메타데이터로 강화",
            "merge_peers": "인접한 유사한 청크를 병합",
            "hierarchical_chunking": "문서 구조를 고려한 계층적 청킹"
        },
        "supported_operations": [
            "chunk(dl_doc): 문서를 청크로 분할",
            "contextualize(chunk): 청크의 컨텍스트 강화"
        ]
    }

if __name__ == "__main__":
    import uvicorn
    logger.info("=" * 80)
    logger.info("🚀 Docling RAG 서버 시작 중...")
    logger.info(f"📍 엔드포인트: http://0.0.0.0:10002")
    logger.info(f"📚 문서: http://0.0.0.0:10002/docs")
    logger.info(f"🔧 비동기 API: POST /convert-chunked-async")
    logger.info(f"🔍 상태 조회: GET /job/{{job_id}}")
    logger.info("=" * 80)
    uvicorn.run(app, host="0.0.0.0", port=10002, log_level="info")